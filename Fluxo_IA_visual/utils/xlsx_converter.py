import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from typing import Dict, Any

from .helpers_texto_fluxo import (
    PALABRAS_EFECTIVO, 
    PALABRAS_TRASPASO_ENTRE_CUENTAS,
    PALABRAS_TRASPASO_FINANCIAMIENTO,
    PALABRAS_BMRCASH,
    PALABRAS_EXCLUIDAS
)

def generar_excel_reporte(data_json: Dict[str, Any]) -> bytes:
    wb = Workbook()
    
    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')

    def aplicar_estilo_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    resultados = data_json.get("resultados_individuales", [])

    # ==========================================
    # 1. RESUMEN POR CUENTA
    # ==========================================
    ws1 = wb.active
    ws1.title = "Resumen por Cuenta"
    ws1.append([
        "Mes", "Cuenta", "Moneda", "Depósitos", "Cargos", "TPV Bruto", 
        "Financiamientos", "Efectivo", "Traspaso entre cuentas", "BMR CASH"
    ])
    
    for res in resultados:
        ia = res.get("AnalisisIA") or {}
        if not ia: continue
        
        periodo = ia.get("periodo_fin") or ia.get("periodo_inicio") or "Desc."
        banco = ia.get("banco", "BANCO")
        clabe = str(ia.get("clabe_interbancaria") or "")
        cuenta_str = f"{banco}-{clabe[-4:]}" if len(clabe) >= 4 else banco
        
        ws1.append([
            periodo, cuenta_str,
            ia.get("tipo_moneda", "MXN"),
            ia.get("depositos", 0.0),
            ia.get("cargos", 0.0),
            ia.get("entradas_TPV_bruto", 0.0),
            ia.get("total_entradas_financiamiento", 0.0),
            ia.get("depositos_en_efectivo", 0.0),
            ia.get("traspaso_entre_cuentas", 0.0),
            ia.get("entradas_bmrcash", 0.0)
        ])

    aplicar_estilo_header(ws1)
    ws1.column_dimensions['B'].width = 25
    for row in ws1.iter_rows(min_row=2, min_col=3, max_col=10):
        for cell in row: cell.style = currency_style

    # ==========================================
    # 2. RESUMEN PORTADAS
    # ==========================================
    ws2 = wb.create_sheet("Resumen Portadas")
    ws2.append([
        "Banco", "RFC", "Cliente", "CLABE / Cuenta", 
        "Periodo Inicio", "Periodo Fin", 
        "Depósitos", "Cargos", "Saldo Promedio", "Comisiones"
    ])
    
    for res in resultados:
        ia = res.get("AnalisisIA") or {}
        clabe_segura = str(ia.get("clabe_interbancaria") or "")
        ws2.append([
            ia.get("banco", "Desconocido"), ia.get("rfc", ""), ia.get("nombre_cliente", ""),
            clabe_segura, ia.get("periodo_inicio", ""), ia.get("periodo_fin", ""),
            ia.get("depositos", 0.0), ia.get("cargos", 0.0),
            ia.get("saldo_promedio", 0.0), ia.get("comisiones", 0.0)
        ])
    
    aplicar_estilo_header(ws2)
    for row in ws2.iter_rows(min_row=2, min_col=7, max_col=10):
        for cell in row: cell.style = currency_style

    # ==========================================
    # HELPER GENERADOR DE HOJAS
    # ==========================================
    def crear_hoja_detalle(nombre_hoja, criterio_filtro_func):
        ws = wb.create_sheet(nombre_hoja)
        ws.append(["Banco", "Fecha", "Descripción", "Monto", "Tipo", "Categoría (IA)"])
        
        for res in resultados:
            ia = res.get("AnalisisIA") or {}
            banco_nom = ia.get("banco", "Desconocido")
            detalle = res.get("DetalleTransacciones", {})
            transacciones = detalle.get("transacciones", [])
            
            if isinstance(transacciones, list):
                for tx in transacciones:
                    desc = str(tx.get("descripcion", "")).lower()
                    tipo = str(tx.get("tipo", "")).lower()
                    # AHORA EXTRAEMOS LA CATEGORÍA PARA EL FILTRO
                    cat = str(tx.get("categoria", "GENERAL")).upper()
                    
                    # --- FILTRADO ---
                    # Pasamos (Descripción, Tipo, Categoría) al filtro
                    if criterio_filtro_func(desc, tipo, cat):
                        try:
                            monto_val = float(str(tx.get("monto", "0")).replace(",", ""))
                        except: monto_val = 0.0

                        ws.append([
                            banco_nom, tx.get("fecha", ""), tx.get("descripcion", ""),
                            monto_val, tx.get("tipo", ""), cat
                        ])
        
        aplicar_estilo_header(ws)
        ws.column_dimensions['C'].width = 60
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row: cell.style = currency_style

    # ==========================================
    # DEFINICIÓN DE FILTROS
    # ==========================================

    def es_excluido(desc):
        return any(p in desc for p in PALABRAS_EXCLUIDAS)

    # 3. TODOS LOS MOVIMIENTOS
    # Nota: Agregamos 'c' (categoría) al lambda aunque no lo usemos, para cumplir con el helper
    crear_hoja_detalle("Todos los Movimientos", lambda d, t, c: not es_excluido(d))

    # 4. TRANSACCIONES TPV (Lógica Estricta Final)
    def filtro_tpv_estricto(desc, tipo, cat):
        # 1. Filtro de basura (Excluidos)
        if es_excluido(desc): return False
        
        # 2. Solo Abonos
        if "abono" not in tipo and "depósito" not in tipo: return False

        # 3. REQUERIMIENTO CLAVE: Categoría NO puede ser GENERAL
        # Como arreglamos el worker, ahora las TPV vendrán etiquetadas como "TPV"
        if cat == "GENERAL": return False

        # 4. Filtro negativo (Redundancia de seguridad)
        if any(p in desc for p in PALABRAS_EFECTIVO): return False
        if any(p in desc for p in PALABRAS_TRASPASO_ENTRE_CUENTAS): return False
        if any(p in desc for p in PALABRAS_TRASPASO_FINANCIAMIENTO): return False
        if any(p in desc for p in PALABRAS_BMRCASH): return False
        
        return True
    
    crear_hoja_detalle("Transacciones TPV", filtro_tpv_estricto)

    # 5. EFECTIVO
    crear_hoja_detalle("Efectivo", lambda d, t, c: not es_excluido(d) and any(p in d for p in PALABRAS_EFECTIVO))

    # 6. FINANCIAMIENTOS
    crear_hoja_detalle("Financiamientos", lambda d, t, c: not es_excluido(d) and any(p in d for p in PALABRAS_TRASPASO_FINANCIAMIENTO))

    # 7. TRASPASO ENTRE CUENTAS
    crear_hoja_detalle("Traspaso entre Cuentas", lambda d, t, c: not es_excluido(d) and any(p in d for p in PALABRAS_TRASPASO_ENTRE_CUENTAS))

    # 8. BMRCASH
    crear_hoja_detalle("BMRCASH", lambda d, t, c: not es_excluido(d) and any(p in d for p in PALABRAS_BMRCASH))

    # ==========================================
    # 9. RESUMEN GENERAL
    # ==========================================
    ws_final = wb.create_sheet("Resumen General")
    ws_final.append(["Métrica", "Valor"])
    
    total_dep = data_json.get("total_depositos", 0.0)
    es_mayor = "SÍ" if data_json.get("es_mayor_a_250") else "NO"
    
    ws_final.append(["Total Depósitos Calculados", total_dep])
    ws_final.append(["¿Es Mayor a 250k?", es_mayor])
    ws_final.append(["Documentos Procesados", len(resultados)])
    
    ws_final['B2'].style = currency_style
    aplicar_estilo_header(ws_final)
    ws_final.column_dimensions['A'].width = 30

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()